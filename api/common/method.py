# @Project: SCB22
# @Auth ： 柠檬班-土豆
# @Time ： 2021/7/16 15:14
# @E-mail ：121313927@qq.com
# @Company：湖南省零檬信息技术有限公司
# @Site: http://www.lemonban.com
# @Forum: http://testingpai.com


import requests
from openpyxl import load_workbook  # 加载load_workbook 函数


# 读取测试用例
def read_case(filename, sheetname):
    wb = load_workbook(filename)  # 打开excel文件
    sh = wb[sheetname]  # 打开某个表单
    max_row = sh.max_row  # 获取总行数
    case_list = []  # 空列表
    for i in range(2, max_row + 1):
        case_dict = dict(
            case_id=sh.cell(row=i, column=1).value,  # 获取case_id
            url=sh.cell(row=i, column=5).value,  # 获取url值
            data=sh.cell(row=i, column=6).value,  # 获取data
            expect=sh.cell(row=i, column=7).value  # 获取expect期望
        )
        case_list.append(case_dict)  # 每次循环，把生成dict追加到list列表中
    return case_list


# 写入测试结果
def write_result(filename, sheetname, row, column,final_result):
    wb = load_workbook(filename)  # 打开excel文件
    sh = wb[sheetname]  # 打开某个表单
    sh.cell(row=row, column=column).value = final_result # 写入数据
    wb.save(filename)  # 保存

# 发送接口请求
def api_fun(url,data):
    header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url,json=data,headers=header).json()
    return res   # 设置返回值